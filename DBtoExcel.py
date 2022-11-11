import pyodbc
import openpyxl as xl

wb = xl.Workbook() # 產生一個工作簿物件
ws1 = wb.create_sheet("系所")  # 產生一張工作表
ws2 = wb.create_sheet("教師")
ws3 = wb.create_sheet("課程")
ws4 = wb.create_sheet("學生")
ws5 = wb.create_sheet("選課單")

##########################################################################

conn = pyodbc.connect(r'Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=D:\Python\DB and Web crawler\教材\資料庫與Python\CH09C2.accdb')
#  系所  教師  課程  學生  選課單
SQL = '''\
SELECT * FROM 系所
'''
cur = conn.cursor()
cur.execute(SQL)
desc = cur.description
# (('科目代碼', <class 'str'>, None, 4, 4, 0, True), ('科目名稱', <class 'str'>, None, 20, 20, 0, True),...)
list1 = cur.fetchall()
# [('A01', '國文', 4, True), ('A02', '應用英文', 3, True)...]
cur.close()
conn.close()

for n,item in enumerate(desc) : # 寫入欄位名稱
    ws1.cell(row=1, column=n+1, value=item[0])

for n,record in enumerate(list1) :  # 寫入所有資料，大迴圈拿出一筆筆的資料
    for m,item in enumerate(record) :  # 小迴圈拿出一個個的元素
        ws1.cell(row=n+2, column=m+1, value=item)

##########################################################################

conn = pyodbc.connect(r'Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=D:\Python\DB and Web crawler\教材\資料庫與Python\CH09C2.accdb')
#  系所  教師  課程  學生  選課單
SQL = '''\
SELECT * FROM 教師
'''
cur = conn.cursor()
cur.execute(SQL)
desc = cur.description
list1 = cur.fetchall()
cur.close()
conn.close()

for n,item in enumerate(desc) : # 寫入欄位名稱
    ws2.cell(row=1, column=n+1, value=item[0])

for n,record in enumerate(list1) :  # 寫入所有資料，大迴圈拿出一筆筆的資料
    for m,item in enumerate(record) :  # 小迴圈拿出一個個的元素
        ws2.cell(row=n+2, column=m+1, value=item)

##########################################################################

conn = pyodbc.connect(r'Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=D:\Python\DB and Web crawler\教材\資料庫與Python\CH09C2.accdb')
#  系所  教師  課程  學生  選課單
SQL = '''\
SELECT * FROM 課程
'''
cur = conn.cursor()
cur.execute(SQL)
desc = cur.description
list1 = cur.fetchall()
cur.close()
conn.close()

for n,item in enumerate(desc) : # 寫入欄位名稱
    ws3.cell(row=1, column=n+1, value=item[0])

for n,record in enumerate(list1) :  # 寫入所有資料，大迴圈拿出一筆筆的資料
    for m,item in enumerate(record) :  # 小迴圈拿出一個個的元素
        ws3.cell(row=n+2, column=m+1, value=item)

##########################################################################

conn = pyodbc.connect(r'Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=D:\Python\DB and Web crawler\教材\資料庫與Python\CH09C2.accdb')
#  系所  教師  課程  學生  選課單
SQL = '''\
SELECT * FROM 學生
'''
cur = conn.cursor()
cur.execute(SQL)
desc = cur.description
list1 = cur.fetchall()
cur.close()
conn.close()

for n,item in enumerate(desc) : # 寫入欄位名稱
    ws4.cell(row=1, column=n+1, value=item[0])

for n,record in enumerate(list1) :  # 寫入所有資料，大迴圈拿出一筆筆的資料
    for m,item in enumerate(record) :  # 小迴圈拿出一個個的元素
        ws4.cell(row=n+2, column=m+1, value=item)

##########################################################################

conn = pyodbc.connect(r'Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=D:\Python\DB and Web crawler\教材\資料庫與Python\CH09C2.accdb')
#  系所  教師  課程  學生  選課單
SQL = '''\
SELECT * FROM 選課單
'''
cur = conn.cursor()
cur.execute(SQL)
desc = cur.description
list1 = cur.fetchall()
cur.close()
conn.close()

for n,item in enumerate(desc) : # 寫入欄位名稱
    ws5.cell(row=1, column=n+1, value=item[0])

for n,record in enumerate(list1) :  # 寫入所有資料，大迴圈拿出一筆筆的資料
    for m,item in enumerate(record) :  # 小迴圈拿出一個個的元素
        ws5.cell(row=n+2, column=m+1, value=item)

###################################################

wb.save("校園.xlsx")
