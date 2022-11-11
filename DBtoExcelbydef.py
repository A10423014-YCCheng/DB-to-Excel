import pyodbc
import openpyxl as xl

def DBtoExcel(tablelist, dbpath, savepath):
    wb = xl.Workbook()
    conn = pyodbc.connect(r'Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=' + dbpath)
    for table in tablelist:
        ws1 = wb.create_sheet(table)
        SQL = 'SELECT * FROM '
        SQL += table
        cur = conn.cursor()
        cur.execute(SQL)
        desc = cur.description
        list1 = cur.fetchall()

        for n, item in enumerate(desc):  # 寫入欄位名稱
            ws1.cell(row=1, column=n + 1, value=item[0])

        for n, record in enumerate(list1):  # 寫入所有資料，大迴圈拿出一筆筆的資料
            for m, item in enumerate(record):  # 小迴圈拿出一個個的元素
                ws1.cell(row=n + 2, column=m + 1, value=item)

    cur.close()
    conn.close()
    wb.save(savepath)

db_l = ['系所','教師','課程','學生','選課單']
dbp = r'D:\Python\DB and Web crawler\教材\資料庫與Python\CH09C2.accdb'
save = r'D:\Python\Program\DB and Web crawler\20221102\homework2.xlsx'
DBtoExcel(db_l, dbp, save)